import os
import logging
import numpy as np
from datetime import datetime, timezone
from db.session import SessionLocal

from db.models import (
    Portfolio, PortfolioPosition, Security, CurveCalibration, ReportGeneration,
)
from quant_core.conventions import get_settlement_date, calculate_accrued_interest
from quant_core.cashflow import generate_cashflows
from quant_core.bootstrap import bootstrap_zero_curve
from quant_core.pricing import calculate_dirty_price, calculate_clean_price, calculate_ytm
from quant_core.risk import (
    calculate_macaulay_duration, calculate_modified_duration, calculate_dv01, calculate_convexity,
)
from quant_core.scenario import apply_scenario_shocks
from quant_core.krd import calculate_key_rate_durations, DEFAULT_KEY_TENORS
from quant_core.nss import nss_yield

logger = logging.getLogger("report_generator")

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")


def _build_zero_curve(params: dict):
    fn = lambda t: nss_yield(t, params["beta0"], params["beta1"], params["beta2"],
                             params["beta3"], params["tau1"], params["tau2"])
    return bootstrap_zero_curve(fn, 40.0, 0.5)


def _compute_position(sec, face_value, base_zc, shocked_zc, sd):
    issue_date = sec.issue_date
    maturity_date = sec.maturity_date
    coupon_rate = float(sec.coupon_rate)
    freq = int(sec.coupon_frequency or 2)

    cfs = generate_cashflows(issue_date, maturity_date, coupon_rate, freq, 100.0)

    base_accrued = calculate_accrued_interest(sd, issue_date, maturity_date, coupon_rate, freq, 100.0)
    base_dirty = calculate_dirty_price(sd, cfs, base_zc)
    base_clean = calculate_clean_price(sd, issue_date, maturity_date, coupon_rate, cfs, base_zc, freq, 100.0)
    ytm = calculate_ytm(sd, cfs, base_dirty, freq)
    mac_dur = calculate_macaulay_duration(sd, cfs, ytm, freq)
    mod_dur = calculate_modified_duration(mac_dur, ytm, freq)
    dv01 = calculate_dv01(sd, cfs, base_zc)
    conv = calculate_convexity(sd, cfs, base_zc)
    krd = calculate_key_rate_durations(sd, cfs, base_zc, DEFAULT_KEY_TENORS)

    shocked_dirty = calculate_dirty_price(sd, cfs, shocked_zc)
    shocked_clean = calculate_clean_price(sd, issue_date, maturity_date, coupon_rate, cfs, shocked_zc, freq, 100.0)
    pnl = (shocked_dirty - base_dirty) * (face_value / 100.0)

    return {
        "isin": sec.isin,
        "name": sec.security_name,
        "face_value": face_value,
        "base_clean": base_clean,
        "base_dirty": base_dirty,
        "accrued": base_accrued,
        "ytm": ytm,
        "mac_dur": mac_dur,
        "mod_dur": mod_dur,
        "dv01": dv01 * (face_value / 100.0),
        "convexity": conv,
        "krd": krd,
        "shocked_dirty": shocked_dirty,
        "shocked_clean": shocked_clean,
        "pnl": pnl,
    }


def _aggregate(positions):
    total_base_dirty = sum(p["base_dirty"] * (p["face_value"] / 100.0) for p in positions)
    total_base_clean = sum(p["base_clean"] * (p["face_value"] / 100.0) for p in positions)
    total_shocked_dirty = sum(p["shocked_dirty"] * (p["face_value"] / 100.0) for p in positions)
    total_pnl = sum(p["pnl"] for p in positions)
    total_dv01 = sum(p["dv01"] for p in positions)

    if total_base_dirty > 0:
        port_mac = sum(p["mac_dur"] * p["base_dirty"] * (p["face_value"] / 100.0) for p in positions) / total_base_dirty
        port_mod = sum(p["mod_dur"] * p["base_dirty"] * (p["face_value"] / 100.0) for p in positions) / total_base_dirty
        port_conv = sum(p["convexity"] * p["base_dirty"] * (p["face_value"] / 100.0) for p in positions) / total_base_dirty
    else:
        port_mac = port_mod = port_conv = 0.0

    port_krd = [0.0] * len(DEFAULT_KEY_TENORS)
    for p in positions:
        for k in range(len(DEFAULT_KEY_TENORS)):
            port_krd[k] += p["krd"][k] * (p["face_value"] / 100.0)

    return {
        "total_base_dirty": total_base_dirty,
        "total_base_clean": total_base_clean,
        "total_shocked_dirty": total_shocked_dirty,
        "total_pnl": total_pnl,
        "total_dv01": total_dv01,
        "port_mac_dur": port_mac,
        "port_mod_dur": port_mod,
        "port_convexity": port_conv,
        "port_krd": port_krd,
    }


def generate_report(report_id: str):
    """Background task: re-derives portfolio/scenario results server-side and renders document."""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    db = SessionLocal()

    try:
        from datetime import timedelta
        stale_cutoff = datetime.now(timezone.utc) - timedelta(minutes=5)
        stale = db.query(ReportGeneration).filter(
            ReportGeneration.status == "processing",
            ReportGeneration.created_at < stale_cutoff,
        ).all()
        for s in stale:
            s.status = "failed"
            s.error_message = "Timed out"
        if stale:
            db.commit()

        rec = db.query(ReportGeneration).filter(ReportGeneration.id == report_id).first()
        if not rec:
            logger.error(f"Report record {report_id} not found")
            return
        if rec.status != "processing":
            logger.info(f"Report {report_id} already {rec.status}, skipping")
            return

        # Load portfolio
        portfolio = db.query(Portfolio).filter(Portfolio.id == rec.portfolio_id).first()
        if not portfolio:
            raise ValueError(f"Portfolio {rec.portfolio_id} not found")

        positions_db = (
            db.query(PortfolioPosition, Security)
            .join(Security, PortfolioPosition.security_id == Security.id)
            .filter(PortfolioPosition.portfolio_id == rec.portfolio_id)
            .all()
        )
        if not positions_db:
            raise ValueError("Portfolio has no positions")

        # Load latest curve calibration
        cal = (
            db.query(CurveCalibration)
            .filter(CurveCalibration.is_active == True)
            .order_by(CurveCalibration.curve_date.desc())
            .first()
        )
        if not cal:
            raise ValueError("No active curve calibration found")

        base_params = {
            "beta0": float(cal.beta0), "beta1": float(cal.beta1),
            "beta2": float(cal.beta2), "beta3": float(cal.beta3),
            "tau1": float(cal.tau1), "tau2": float(cal.tau2),
        }
        base_zc = _build_zero_curve(base_params)
        curve_date = cal.curve_date
        sd = get_settlement_date(curve_date)

        scenarios = rec.scenario_config if isinstance(rec.scenario_config, list) else [rec.scenario_config]
        scenario_results = []

        for sc in scenarios:
            shocks = {
                "parallel_shift": sc.get("parallel_shift", 0.0),
                "slope_shock": sc.get("slope_shock", 0.0),
                "curvature1_shock": sc.get("curvature1_shock", 0.0),
                "curvature2_shock": sc.get("curvature2_shock", 0.0),
                "twist_shock": sc.get("twist_shock", 0.0),
                "twist_pivot": sc.get("twist_pivot", 5.0),
            }
            shocked_params = apply_scenario_shocks(base_params, **shocks)
            shocked_zc = _build_zero_curve(shocked_params)

            pos_results = []
            for pp, sec in positions_db:
                pos_results.append(_compute_position(sec, float(pp.face_value_held), base_zc, shocked_zc, sd))

            agg = _aggregate(pos_results)
            scenario_results.append({
                "name": sc.get("name", "Scenario"),
                "shocks": shocks,
                "positions": pos_results,
                "summary": agg,
                "shocked_params": shocked_params,
            })

        # Build zero curve data for report
        base_zc_data = []
        for t in np.arange(0.5, 30.5, 0.5):
            base_zc_data.append({"tenor": t, "rate": base_zc.get_zero_rate(t), "df": base_zc.get_discount_factor(t)})

        ext = "pdf" if rec.format == "pdf" else "xlsx"
        file_path = os.path.join(REPORTS_DIR, f"{report_id}.{ext}")

        if rec.format == "pdf":
            _render_pdf(file_path, portfolio.portfolio_name, curve_date, scenario_results, base_zc_data)
        else:
            _render_xlsx(file_path, portfolio.portfolio_name, curve_date, scenario_results, base_zc_data)

        rec.status = "completed"
        rec.storage_path = file_path
        rec.generated_at = datetime.now(timezone.utc)
        db.commit()
        logger.info(f"Report {report_id} generated: {file_path}")

    except Exception as e:
        logger.error(f"Report {report_id} failed: {e}")
        try:
            if 'rec' in locals() and rec is not None:
                rec.status = "failed"
                rec.error_message = str(e)
                db.commit()
        except Exception:
            logger.error(f"Failed to update report status for {report_id}")
    finally:
        db.close()


def _render_pdf(path, portfolio_name, curve_date, scenario_results, base_zc_data):
    from fpdf import FPDF

    C_BG = (12, 13, 17)
    C_SURFACE = (19, 20, 26)
    C_ELEVATED = (26, 27, 34)
    C_ACCENT = (242, 169, 0)
    C_TEXT = (232, 230, 225)
    C_DIM = (150, 150, 150)
    C_BORDER = (55, 55, 70)
    C_POS = (45, 155, 117)
    C_NEG = (235, 115, 85)
    SITE_URL = "https://bondfactor.vercel.app"
    LINKEDIN_URL = "https://linkedin.com/in/sourabh-pradhan07/"

    class StyledPDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(*C_BG)
            self.rect(0, 0, self.w, self.h, "F")
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(*C_ACCENT)
            self.cell(0, 8, "BondFactor", new_x="LMARGIN", new_y="NEXT", link=SITE_URL)
            self.set_draw_color(*C_ACCENT)
            self.set_line_width(0.5)
            self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_draw_color(*C_BORDER)
            self.set_line_width(0.25)
            self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
            self.ln(3)
            half = (self.w - self.l_margin - self.r_margin) / 2
            self.set_font("Helvetica", "", 6.5)
            self.set_text_color(*C_DIM)
            self.cell(half, 5, "Made by Sourabh", link=LINKEDIN_URL)
            self.cell(half, 5, f"Page {self.page_no()}/{{nb}}", align="R", new_x="LMARGIN", new_y="NEXT")
            self.cell(half, 5, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
            self.cell(half, 5, "bondfactor.vercel.app", align="R", link=SITE_URL, new_x="LMARGIN", new_y="NEXT")

    pdf = StyledPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(left=15, top=20, right=15)
    W = pdf.w - pdf.l_margin - pdf.r_margin

    def _color(val):
        return C_POS if val > 0 else C_NEG if val < 0 else C_TEXT

    def _fill_bg():
        pdf.set_fill_color(*C_BG)
        pdf.rect(0, 0, pdf.w, pdf.h, "F")

    def _section(title):
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*C_ACCENT)
        pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(*C_BORDER)
        pdf.set_line_width(0.5)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + 80, pdf.get_y())
        pdf.ln(4)

    def _table_header(headers, widths):
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*C_ACCENT)
        pdf.set_fill_color(*C_SURFACE)
        for h, w in zip(headers, widths):
            pdf.cell(w, 7, h, fill=True)
        pdf.ln()

    def _table_row(values, widths, aligns, row_idx, fmts=None, color_cols=None):
        pdf.set_font("Helvetica", "", 7)
        pdf.set_fill_color(*(C_ELEVATED if row_idx % 2 else C_BG))
        for i, (v, w, a) in enumerate(zip(values, widths, aligns)):
            txt = fmts[i](v) if fmts else str(v)
            if color_cols and i in color_cols and isinstance(v, (int, float)):
                pdf.set_text_color(*_color(v))
            else:
                pdf.set_text_color(*C_TEXT)
            pdf.cell(w, 6, txt, align=a, fill=True)
        pdf.ln()

    # ── title page ───────────────────────────────────────────────────────
    pdf.add_page()
    _fill_bg()
    pdf.ln(50)
    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(*C_ACCENT)
    pdf.cell(0, 14, "BondFactor", align="C", new_x="LMARGIN", new_y="NEXT", link=SITE_URL)
    pdf.set_draw_color(*C_ACCENT)
    pdf.set_line_width(1)
    pdf.line(pdf.w / 2 - 30, pdf.get_y(), pdf.w / 2 + 30, pdf.get_y())
    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(*C_TEXT)
    pdf.cell(0, 10, "RISK REPORT", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(18)

    cx = pdf.l_margin + (W - 120) / 2
    pdf.set_fill_color(*C_SURFACE)
    pdf.set_draw_color(*C_BORDER)
    pdf.set_line_width(0.5)
    pdf.rect(cx, pdf.get_y(), 120, 34, "DF")
    pdf.ln(5)
    gen_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    for label, val in [("Portfolio", portfolio_name), ("Curve Date", str(curve_date)),
                       ("Scenarios", str(len(scenario_results))), ("Generated", gen_time)]:
        pdf.set_x(cx + 5)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*C_DIM)
        pdf.cell(40, 7, label)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*C_TEXT)
        pdf.cell(70, 7, val, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(12)

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*C_DIM)
    pdf.cell(0, 5, "This report is generated by the BondFactor fixed-income risk engine. All values are server-side repriced.", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_text_color(*C_ACCENT)
    pdf.cell(0, 5, "bondfactor.vercel.app", align="C", link=SITE_URL, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_text_color(*C_DIM)
    pdf.cell(0, 5, "Made by Sourabh", align="C", link=LINKEDIN_URL, new_x="LMARGIN", new_y="NEXT")

    # ── scenario pages ───────────────────────────────────────────────────
    for sc in scenario_results:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(*C_ACCENT)
        pdf.cell(0, 8, f"SCENARIO: {sc['name']}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        s = sc["shocks"]
        pdf.set_fill_color(*C_SURFACE)
        pdf.set_draw_color(*C_BORDER)
        pdf.set_line_width(0.5)
        pdf.rect(pdf.l_margin, pdf.get_y(), W, 8, "DF")
        pdf.ln(1)
        for txt in [f"Parallel: {s['parallel_shift']:+.2f}%", f"Slope: {s['slope_shock']:+.2f}%",
                    f"Curv1: {s['curvature1_shock']:+.2f}%", f"Curv2: {s['curvature2_shock']:+.2f}%",
                    f"Twist: {s['twist_shock']:+.2f}%", f"Pivot: {s['twist_pivot']:.1f}Y"]:
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(*C_DIM)
            pdf.cell(W / 6, 6, txt)
        pdf.ln(10)

        agg = sc["summary"]
        _section("PORTFOLIO RISK SUMMARY")
        metrics = [
            ("TOTAL BASE DIRTY", f"Rs. {agg['total_base_dirty']:,.2f}"),
            ("CLEAN VALUE", f"Rs. {agg['total_base_clean']:,.2f}"),
            ("SHOCKED VALUE", f"Rs. {agg['total_shocked_dirty']:,.2f}"),
            ("SCENARIO P&L", f"Rs. {agg['total_pnl']:+,.2f}"),
            ("MOD. DURATION", f"{agg['port_mod_dur']:.4f} Y"),
            ("TOTAL DV01", f"Rs. {agg['total_dv01']:,.2f}"),
            ("CONVEXITY", f"{agg['port_convexity']:.4f}"),
        ]
        card_w = (W - 6) / 4
        card_h = 18
        y0 = pdf.get_y()
        for i, (label, val) in enumerate(metrics):
            col, row = i % 4, i // 4
            cx_card = pdf.l_margin + col * (card_w + 2)
            cy_card = y0 + row * (card_h + 2)
            pdf.set_fill_color(*C_SURFACE)
            pdf.set_draw_color(*C_BORDER)
            pdf.set_line_width(0.3)
            pdf.rect(cx_card, cy_card, card_w, card_h, "DF")
            pdf.set_xy(cx_card + 3, cy_card + 2)
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(*C_DIM)
            pdf.cell(card_w - 4, 4, label)
            pdf.set_xy(cx_card + 3, cy_card + 8)
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(*_color(agg["total_pnl"]) if "P&L" in label else C_TEXT)
            pdf.cell(card_w - 4, 6, val)
        pdf.set_y(y0 + 2 * (card_h + 2) + 5)

        _section("ZERO CURVE (BASE)")
        _zc_tenors = {0.5, 1, 2, 3, 5, 7, 10, 15, 20, 25, 30}
        _zc_filtered = [pt for pt in base_zc_data if pt["tenor"] in _zc_tenors]
        zc_widths = [40, 50, 50]
        _table_header(["Tenor", "Zero Rate (%)", "Discount Factor"], zc_widths)
        for i, pt in enumerate(_zc_filtered):
            _table_row(
                [pt["tenor"], pt["rate"], pt["df"]], zc_widths,
                ["L", "R", "R"], i,
                [lambda v: f"{v:.1f}Y", lambda v: f"{v:.4f}", lambda v: f"{v:.6f}"],
            )
        pdf.ln(4)

        _section("PORTFOLIO POSITIONS")
        pos_widths = [30, 22, 20, 20, 15, 15, 20, 22]
        _table_header(["ISIN", "Face Value", "Clean", "Dirty", "YTM%", "ModDur", "DV01", "P&L"], pos_widths)
        for i, p in enumerate(sc["positions"]):
            _table_row(
                [p["isin"], p["face_value"], p["base_clean"], p["base_dirty"],
                 p["ytm"], p["mod_dur"], p["dv01"], p["pnl"]],
                pos_widths, ["L", "R", "R", "R", "R", "R", "R", "R"], i,
                [lambda v: str(v)[:15], lambda v: f"{v:,.0f}", lambda v: f"{v:.4f}",
                 lambda v: f"{v:.4f}", lambda v: f"{v:.3f}", lambda v: f"{v:.3f}",
                 lambda v: f"{v:,.0f}", lambda v: f"{v:+,.2f}"],
                color_cols={7},
            )
        pdf.ln(4)

        _section("KEY RATE DURATIONS")
        tenor_labels = [f"{t}Y" if t >= 1 else f"{int(t * 12)}M" for t in DEFAULT_KEY_TENORS]
        krd_w = W / len(tenor_labels)
        _table_header(tenor_labels, [krd_w] * len(tenor_labels))
        _table_row(
            agg["port_krd"], [krd_w] * len(tenor_labels),
            ["R"] * len(tenor_labels), 0,
            [lambda v: f"{v:.3f}"] * len(tenor_labels),
        )

    pdf.output(path)


def _render_xlsx(path, portfolio_name, curve_date, scenario_results, base_zc_data):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference, BarChart
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="13141A", end_color="13141A", fill_type="solid")
    lbl_font = Font(bold=True, size=10)
    green_font = Font(color="2D9B75")
    red_font = Font(color="EB7355")
    num_fmt_money = '#,##0.00'
    num_fmt_int = '#,##0'
    num_fmt_4 = '0.0000'
    num_fmt_6 = '0.000000'
    num_fmt_3 = '0.000'

    def _set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Summary sheet ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["BondFactor Risk Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Portfolio: {portfolio_name}"])
    ws.append([f"Curve Date: {curve_date}"])
    ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.append([])
    _set_col_widths(ws, [30, 20])

    for sc in scenario_results:
        ws.append([f"Scenario: {sc['name']}"])
        ws.cell(ws.max_row, 1).font = lbl_font
        agg = sc["summary"]
        for label, val, fmt in [
            ("Total Base Dirty Value", agg["total_base_dirty"], num_fmt_money),
            ("Total Base Clean Value", agg["total_base_clean"], num_fmt_money),
            ("Total Shocked Dirty Value", agg["total_shocked_dirty"], num_fmt_money),
            ("Scenario P&L", agg["total_pnl"], num_fmt_money),
            ("Modified Duration", agg["port_mod_dur"], num_fmt_4),
            ("Total DV01", agg["total_dv01"], num_fmt_money),
            ("Convexity", agg["port_convexity"], num_fmt_4),
        ]:
            ws.append([label, val])
            r = ws.max_row
            ws.cell(r, 1).font = lbl_font
            ws.cell(r, 2).number_format = fmt
            if "P&L" in label:
                ws.cell(r, 2).font = green_font if val >= 0 else red_font
        ws.append([])

    # ── Zero Curve sheet ─────────────────────────────────────────────────
    ws_zc = wb.create_sheet("Zero Curve")
    ws_zc.append(["Tenor (Y)", "Zero Rate (%)", "Discount Factor"])
    _style_header(ws_zc, 1, 3)
    _set_col_widths(ws_zc, [12, 16, 16])
    for pt in base_zc_data:
        ws_zc.append([pt["tenor"], pt["rate"], pt["df"]])
        r = ws_zc.max_row
        ws_zc.cell(r, 2).number_format = num_fmt_4
        ws_zc.cell(r, 3).number_format = num_fmt_6
    ws_zc.freeze_panes = "A2"

    chart = LineChart()
    chart.title = "Base Zero Curve"
    chart.y_axis.title = "Zero Rate (%)"
    chart.x_axis.title = "Maturity (Years)"
    chart.width = 20
    chart.height = 12
    chart.style = 10
    data_ref = Reference(ws_zc, min_col=2, min_row=1, max_row=len(base_zc_data) + 1)
    cats = Reference(ws_zc, min_col=1, min_row=2, max_row=len(base_zc_data) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws_zc.add_chart(chart, "E2")

    # ── Per-scenario sheets ──────────────────────────────────────────────
    pos_headers = ["ISIN", "Security", "Face Value", "Base Clean", "Base Dirty", "Accrued",
                   "YTM%", "Mod Dur", "DV01", "Convexity", "Shocked Dirty", "P&L"]
    pos_widths = [15, 30, 14, 14, 14, 12, 10, 10, 14, 12, 14, 14]
    pos_fmts = [None, None, num_fmt_int, num_fmt_4, num_fmt_4, num_fmt_4,
                num_fmt_4, num_fmt_4, num_fmt_money, num_fmt_4, num_fmt_4, num_fmt_money]

    for sc_idx, sc in enumerate(scenario_results):
        ws_sc = wb.create_sheet(f"Scenario {sc_idx + 1}")
        s = sc["shocks"]
        ws_sc.append([f"Scenario: {sc['name']}"])
        ws_sc.cell(1, 1).font = Font(bold=True, size=12)
        ws_sc.append([f"Parallel: {s['parallel_shift']:+.2f}%  Slope: {s['slope_shock']:+.2f}%  "
                      f"Curv1: {s['curvature1_shock']:+.2f}%  Curv2: {s['curvature2_shock']:+.2f}%  "
                      f"Twist: {s['twist_shock']:+.2f}%  Pivot: {s['twist_pivot']:.1f}Y"])
        ws_sc.append([])

        header_row = ws_sc.max_row + 1
        ws_sc.append(pos_headers)
        _style_header(ws_sc, header_row, len(pos_headers))
        _set_col_widths(ws_sc, pos_widths)

        for p in sc["positions"]:
            ws_sc.append([p["isin"], p["name"], p["face_value"], p["base_clean"], p["base_dirty"],
                          p["accrued"], p["ytm"], p["mod_dur"], p["dv01"], p["convexity"],
                          p["shocked_dirty"], p["pnl"]])
            r = ws_sc.max_row
            for c, fmt in enumerate(pos_fmts, 1):
                if fmt:
                    ws_sc.cell(r, c).number_format = fmt
            pnl_cell = ws_sc.cell(r, 12)
            pnl_cell.font = green_font if p["pnl"] >= 0 else red_font

        ws_sc.freeze_panes = f"A{header_row + 1}"

        # ── KRD sheet ────────────────────────────────────────────────────
        ws_krd = wb.create_sheet(f"KRD {sc_idx + 1}")
        tenor_labels = [f"{t}Y" if t >= 1 else f"{int(t * 12)}M" for t in DEFAULT_KEY_TENORS]
        ws_krd.append(["Tenor"] + tenor_labels)
        _style_header(ws_krd, 1, len(tenor_labels) + 1)
        ws_krd.append(["Portfolio KRD"] + sc["summary"]["port_krd"])
        for c in range(2, len(tenor_labels) + 2):
            ws_krd.cell(2, c).number_format = num_fmt_3
        _set_col_widths(ws_krd, [14] + [10] * len(tenor_labels))

        krd_chart = BarChart()
        krd_chart.title = f"KRD Profile — {sc['name']}"
        krd_chart.y_axis.title = "Sensitivity (₹/bp)"
        krd_chart.width = 20
        krd_chart.height = 12
        krd_chart.style = 10
        krd_data = Reference(ws_krd, min_col=2, max_col=len(tenor_labels) + 1, min_row=2)
        krd_cats = Reference(ws_krd, min_col=2, max_col=len(tenor_labels) + 1, min_row=1)
        krd_chart.add_data(krd_data, from_rows=True, titles_from_data=False)
        krd_chart.set_categories(krd_cats)
        ws_krd.add_chart(krd_chart, "A4")

    wb.save(path)
